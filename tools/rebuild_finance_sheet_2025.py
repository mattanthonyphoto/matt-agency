"""
Rebuild 2025 Google Finance Sheet from Excel source of truth.
Reads all tabs from the Excel file and writes a clean Transactions tab.
"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from datetime import datetime, date
from google_sheets_auth import get_sheets_service

SPREADSHEET_ID = "1TyX4sOIo6CxblxjNM9Yft5h5vZO_BGhQjHlgC0uzQtI"
EXCEL_PATH = "/Users/matthewfernandes/Downloads/2025 Business Finances.xlsx"
SHEET_NAME = "Transactions"

# Excel date serial: days since Dec 30, 1899
EXCEL_EPOCH = date(1899, 12, 30)

def to_serial(d):
    """Convert date/datetime to Excel serial number for Google Sheets."""
    if isinstance(d, datetime):
        d = d.date()
    if isinstance(d, date):
        return (d - EXCEL_EPOCH).days
    return None

def month_serial(month_num):
    """Return serial for the 15th of given month in 2025."""
    return to_serial(date(2025, month_num, 15))

# Tab -> (category, is_business, payment_method)
# Notes:
#  - Office uses "Total" column header, payment = Debit (Sacred Dance e-transfer)
#  - All others use "Amount (Incl. GST)" column header
TAB_MAP = {
    "Office":           ("Rent / Co-working",           True,  "Debit"),
    "Assets":           ("Equipment (CCA)",              True,  "Visa"),
    "Advertising":      ("Advertising & Marketing",      True,  "Visa"),
    "Vehicle expenses": ("Vehicle",                      True,  "Visa"),
    "Office expenses":  ("Software & Subscriptions",     True,  "Visa"),
    "Supplies":         ("Office Supplies (<$500)",      True,  "Visa"),
    "Education":        ("Software & Subscriptions",     True,  "Visa"),
    "Banking":          ("Interest & Bank Charges",      True,  "Visa"),
    "Travel":           ("Travel",                       True,  "Visa"),
    "Insurance":        ("Insurance",                    True,  "Visa"),
    "Interest":         ("Interest & Bank Charges",      True,  "Visa"),
    "Accounting":       ("Professional Fees",            True,  "Visa"),
    "Phone":            ("Telephone & Internet",         True,  "Visa"),
    "Meals":            ("Meals & Entertainment",        True,  "Visa"),
    "Miscellaneous":    ("Other Business",               True,  "Visa"),
    "Entertainment":    ("Entertainment",                False, "Visa"),
    "Groceries":        ("Groceries",                    False, "Visa"),
    "Fitness":          ("Health & Fitness",             False, "Visa"),
    "Personal":         ("Other Personal",               False, "Visa"),
}

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]

def load_excel_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ---- INCOME ----
    income_rows = []
    ws = wb["Income"]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))

    # Row 0: month headers (col B onwards = indices 1..12)
    # Rows 2..13 (index 2..): payment amounts per month
    # Find where data rows end (Monthly Income row)
    data_rows = []
    for row in all_rows[2:]:  # skip header rows
        if row[0] == "Monthly Income":
            break
        data_rows.append(row)

    for row in data_rows:
        for col_idx, month_name in enumerate(MONTHS, start=1):
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and val != 0:
                month_num = col_idx  # 1=Jan...12=Dec
                income_rows.append({
                    "date_serial": month_serial(month_num),
                    "vendor": "Client Revenue",
                    "description": f"{month_name} Payment",
                    "amount": float(val),  # positive
                    "business": True,
                    "category": "Income",
                    "split": "100%",
                    "gst": 0,
                    "receipt": False,
                    "payment": "E-Transfer",
                    "account": "",
                    "notes": "",
                })

    print(f"Income rows: {len(income_rows)}, Total: ${sum(r['amount'] for r in income_rows):.2f}")

    # ---- EXPENSES ----
    expense_rows = []
    for tab_name, (category, is_business, payment) in TAB_MAP.items():
        ws = wb[tab_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        count = 0
        tab_total = 0
        for row in rows:
            # Skip empty rows
            if not any(v is not None for v in row[:3]):
                continue
            raw_date = row[0]
            description = str(row[1]).strip() if row[1] else ""
            amount_raw = row[2]
            notes = str(row[3]).strip() if row[3] else ""

            if amount_raw is None:
                continue

            amount = float(amount_raw)
            if amount == 0:
                continue

            # Convert date
            if isinstance(raw_date, (datetime, date)):
                date_serial = to_serial(raw_date)
            else:
                date_serial = to_serial(date(2025, 1, 15))  # fallback

            # Vendor: extract first line of description
            vendor = description.split("\n")[0].strip() if "\n" in description else description
            full_desc = description.replace("\n", " - ") if "\n" in description else description

            # Split% and GST
            if is_business:
                split_pct = "100%"
                gst = round(amount / 1.05 * 0.05, 4)
            else:
                split_pct = "0%"
                gst = 0

            expense_rows.append({
                "date_serial": date_serial,
                "vendor": vendor,
                "description": full_desc,
                "amount": -amount,  # negative for expenses
                "business": is_business,
                "category": category,
                "split": split_pct,
                "gst": gst,
                "receipt": False,
                "payment": payment,
                "account": "",
                "notes": notes,
                "_tab": tab_name,
            })
            count += 1
            tab_total += amount

        print(f"  {tab_name}: {count} rows, ${tab_total:.2f}")

    # Sort expenses by date
    expense_rows.sort(key=lambda r: r["date_serial"] if r["date_serial"] else 0)

    print(f"\nExpense rows: {len(expense_rows)}, Total: ${sum(-r['amount'] for r in expense_rows):.2f}")

    return income_rows, expense_rows


def build_sheet_data(income_rows, expense_rows):
    """Build the full list of rows to write to the sheet."""
    rows = []

    # Rows 1-3: Title rows — leave blank (we preserve them, write from row 4)
    # We'll start writing at row 4

    # Income total row (row 4)
    income_total = sum(r["amount"] for r in income_rows)
    rows.append(["INCOME", "", "", f"Total: ${income_total:,.2f}", "", "", "", "", "", "", "", ""])

    # Column headers (row 5)
    rows.append(["Date", "Vendor", "Description", "", "Business", "Category", "", "", "Receipt", "Payment", "Account", "Notes"])

    # Income data
    for r in income_rows:
        rows.append([
            r["date_serial"],
            r["vendor"],
            r["description"],
            r["amount"],
            r["business"],
            r["category"],
            r["split"],
            r["gst"],
            r["receipt"],
            r["payment"],
            r["account"],
            r["notes"],
        ])

    # Spacer
    rows.append([""] * 12)

    # Expenses total row
    expense_total = sum(-r["amount"] for r in expense_rows)
    rows.append(["EXPENSES", "", "", f"Total: ${expense_total:,.2f}", "", "", "", "", "", "", "", ""])

    # Column headers again
    rows.append(["Date", "Vendor", "Description", "", "Business", "Category", "", "", "Receipt", "Payment", "Account", "Notes"])

    # Expense data
    for r in expense_rows:
        rows.append([
            r["date_serial"],
            r["vendor"],
            r["description"],
            r["amount"],
            r["business"],
            r["category"],
            r["split"],
            r["gst"],
            r["receipt"],
            r["payment"],
            r["account"],
            r["notes"],
        ])

    return rows


def clear_and_write(service, sheet_data):
    """Clear existing data from row 4 down and write new data."""

    # First get sheet info to find the sheet ID
    meta = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
    sheet_id = None
    for s in meta["sheets"]:
        if s["properties"]["title"] == SHEET_NAME:
            sheet_id = s["properties"]["sheetId"]
            break

    if sheet_id is None:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found!")
        print("Available sheets:", [s["properties"]["title"] for s in meta["sheets"]])
        return None

    print(f"Found sheet '{SHEET_NAME}' with ID {sheet_id}")

    # Clear from row 4 to end (keep rows 1-3 title)
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A4:Z2000"
    ).execute()
    print("Cleared existing data from row 4+")

    # Write the data starting at row 4
    body = {"values": sheet_data}
    result = service.spreadsheets().values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A4",
        valueInputOption="RAW",
        body=body
    ).execute()

    print(f"Written {result.get('updatedRows', 0)} rows, {result.get('updatedColumns', 0)} cols")

    return sheet_id


def apply_formatting(service, sheet_id, income_rows, expense_rows):
    """Apply all formatting in a batch request."""
    requests = []

    # Calculate row positions (0-indexed for API, data starts at row 4 = index 3)
    # Row 4 (index 3): INCOME header
    # Row 5 (index 4): income column headers
    # Rows 6..(5+len(income)): income data  (index 5..4+len(income))
    # Then spacer
    # Then EXPENSES header
    # Then expense column headers
    # Then expense data

    income_header_row = 3   # 0-indexed
    income_col_header_row = 4
    income_data_start = 5
    income_data_end = income_data_start + len(income_rows)  # exclusive
    spacer_row = income_data_end
    expense_header_row = spacer_row + 1
    expense_col_header_row = expense_header_row + 1
    expense_data_start = expense_col_header_row + 1
    expense_data_end = expense_data_start + len(expense_rows)

    def rgb(r, g, b):
        return {"red": r/255, "green": g/255, "blue": b/255}

    # INCOME header - green background
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": income_header_row,
                "endRowIndex": income_header_row + 1,
                "startColumnIndex": 0,
                "endColumnIndex": 12,
            },
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": rgb(52, 168, 83),
                    "textFormat": {"bold": True, "foregroundColor": rgb(255, 255, 255)},
                }
            },
            "fields": "userEnteredFormat(backgroundColor,textFormat)",
        }
    })

    # EXPENSES header - red background
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": expense_header_row,
                "endRowIndex": expense_header_row + 1,
                "startColumnIndex": 0,
                "endColumnIndex": 12,
            },
            "cell": {
                "userEnteredFormat": {
                    "backgroundColor": rgb(234, 67, 53),
                    "textFormat": {"bold": True, "foregroundColor": rgb(255, 255, 255)},
                }
            },
            "fields": "userEnteredFormat(backgroundColor,textFormat)",
        }
    })

    # Column header rows - bold gray
    for col_header_row in [income_col_header_row, expense_col_header_row]:
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": col_header_row,
                    "endRowIndex": col_header_row + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 12,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": rgb(243, 243, 243),
                        "textFormat": {"bold": True},
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat)",
            }
        })

    # Date format for column A (all data rows)
    for data_start, data_end in [(income_data_start, income_data_end), (expense_data_start, expense_data_end)]:
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": data_start,
                    "endRowIndex": data_end,
                    "startColumnIndex": 0,
                    "endColumnIndex": 1,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "DATE", "pattern": "MMM d, yyyy"}
                    }
                },
                "fields": "userEnteredFormat.numberFormat",
            }
        })

    # Amount (col D, index 3) - currency format
    for data_start, data_end in [(income_data_start, income_data_end), (expense_data_start, expense_data_end)]:
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": data_start,
                    "endRowIndex": data_end,
                    "startColumnIndex": 3,
                    "endColumnIndex": 4,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "CURRENCY", "pattern": '"$"#,##0.00'}
                    }
                },
                "fields": "userEnteredFormat.numberFormat",
            }
        })

    # Business checkbox (col E, index 4) - income
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": income_data_start,
                "endRowIndex": income_data_end,
                "startColumnIndex": 4,
                "endColumnIndex": 5,
            },
            "cell": {
                "dataValidation": {
                    "condition": {"type": "BOOLEAN"},
                    "strict": True,
                    "showCustomUi": True,
                }
            },
            "fields": "dataValidation",
        }
    })

    # Business checkbox (col E, index 4) - expenses
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": expense_data_start,
                "endRowIndex": expense_data_end,
                "startColumnIndex": 4,
                "endColumnIndex": 5,
            },
            "cell": {
                "dataValidation": {
                    "condition": {"type": "BOOLEAN"},
                    "strict": True,
                    "showCustomUi": True,
                }
            },
            "fields": "dataValidation",
        }
    })

    # Receipt checkbox (col I, index 8) - income
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": income_data_start,
                "endRowIndex": income_data_end,
                "startColumnIndex": 8,
                "endColumnIndex": 9,
            },
            "cell": {
                "dataValidation": {
                    "condition": {"type": "BOOLEAN"},
                    "strict": True,
                    "showCustomUi": True,
                }
            },
            "fields": "dataValidation",
        }
    })

    # Receipt checkbox (col I, index 8) - expenses
    requests.append({
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": expense_data_start,
                "endRowIndex": expense_data_end,
                "startColumnIndex": 8,
                "endColumnIndex": 9,
            },
            "cell": {
                "dataValidation": {
                    "condition": {"type": "BOOLEAN"},
                    "strict": True,
                    "showCustomUi": True,
                }
            },
            "fields": "dataValidation",
        }
    })

    # Category dropdown (col F, index 5) - all data rows
    categories = [
        "Income",
        "Rent / Co-working",
        "Equipment (CCA)",
        "Advertising & Marketing",
        "Vehicle",
        "Software & Subscriptions",
        "Office Supplies (<$500)",
        "Interest & Bank Charges",
        "Travel",
        "Insurance",
        "Professional Fees",
        "Telephone & Internet",
        "Meals & Entertainment",
        "Other Business",
        "Entertainment",
        "Groceries",
        "Health & Fitness",
        "Other Personal",
    ]
    for data_start, data_end in [(income_data_start, income_data_end), (expense_data_start, expense_data_end)]:
        requests.append({
            "setDataValidation": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": data_start,
                    "endRowIndex": data_end,
                    "startColumnIndex": 5,
                    "endColumnIndex": 6,
                },
                "rule": {
                    "condition": {
                        "type": "ONE_OF_LIST",
                        "values": [{"userEnteredValue": c} for c in categories],
                    },
                    "strict": False,
                    "showCustomUi": True,
                }
            }
        })

    # Blue background for business rows in expenses
    # We need to apply per-row based on business flag
    # Do it in batches - find business vs personal row ranges
    # For expense rows, apply blue tint to business rows
    blue_tint = rgb(235, 245, 255)
    personal_tint = rgb(255, 245, 240)

    for idx, r in enumerate(expense_rows):
        row_index = expense_data_start + idx
        bg = blue_tint if r["business"] else personal_tint
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_index,
                    "endRowIndex": row_index + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 12,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": bg,
                    }
                },
                "fields": "userEnteredFormat.backgroundColor",
            }
        })

    # Income rows - light green tint
    income_tint = rgb(240, 250, 243)
    for idx in range(len(income_rows)):
        row_index = income_data_start + idx
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_index,
                    "endRowIndex": row_index + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": 12,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": income_tint,
                    }
                },
                "fields": "userEnteredFormat.backgroundColor",
            }
        })

    # Amount color: green for income, red for expenses (col D)
    for idx in range(len(income_rows)):
        row_index = income_data_start + idx
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_index,
                    "endRowIndex": row_index + 1,
                    "startColumnIndex": 3,
                    "endColumnIndex": 4,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "CURRENCY", "pattern": '"$"#,##0.00'},
                        "textFormat": {"foregroundColor": rgb(52, 168, 83)},
                    }
                },
                "fields": "userEnteredFormat(numberFormat,textFormat)",
            }
        })

    for idx in range(len(expense_rows)):
        row_index = expense_data_start + idx
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_index,
                    "endRowIndex": row_index + 1,
                    "startColumnIndex": 3,
                    "endColumnIndex": 4,
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {"type": "CURRENCY", "pattern": '"$"#,##0.00'},
                        "textFormat": {"foregroundColor": rgb(234, 67, 53)},
                    }
                },
                "fields": "userEnteredFormat(numberFormat,textFormat)",
            }
        })

    # Send formatting in batches (API limit ~100 requests per call)
    batch_size = 100
    total = len(requests)
    print(f"Sending {total} formatting requests in batches of {batch_size}...")
    for i in range(0, total, batch_size):
        batch = requests[i:i+batch_size]
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": batch}
        ).execute()
        print(f"  Sent batch {i//batch_size + 1}/{(total+batch_size-1)//batch_size}")

    print("Formatting applied.")

    return income_data_start, income_data_end, expense_data_start, expense_data_end


def write_summary(income_rows, expense_rows):
    """Print final summary with row counts and category totals."""
    print("\n" + "="*60)
    print("IMPORT SUMMARY")
    print("="*60)

    income_total = sum(r["amount"] for r in income_rows)
    print(f"\nINCOME: {len(income_rows)} rows")
    print(f"  Total: ${income_total:,.2f}")
    print(f"  Expected: $105,566.26")
    diff = abs(income_total - 105566.26)
    print(f"  Difference: ${diff:.2f} {'OK' if diff < 0.01 else 'WARNING'}")

    print(f"\nEXPENSES: {len(expense_rows)} rows")

    # Category breakdown
    from collections import defaultdict
    by_cat = defaultdict(lambda: {"count": 0, "total": 0})
    business_total = 0
    for r in expense_rows:
        cat = r["category"]
        by_cat[cat]["count"] += 1
        by_cat[cat]["total"] += -r["amount"]
        if r["business"]:
            business_total += -r["amount"]

    print("\n  By Category:")
    for cat, data in sorted(by_cat.items()):
        print(f"    {cat:<30} {data['count']:>4} rows  ${data['total']:>10,.2f}")

    print(f"\n  Business Total: ${business_total:,.2f}")
    print(f"  Expected:       $59,479.34")
    diff = abs(business_total - 59479.34)
    print(f"  Difference:     ${diff:.2f} {'OK' if diff < 1.0 else 'WARNING'}")

    # Key category checks
    print("\n  Key Category Checks:")
    checks = {
        "Rent / Co-working": 9690,
        "Equipment (CCA)": 18967,
        "Advertising & Marketing": 6720,
        "Vehicle": 3696,
    }
    for cat, expected in checks.items():
        actual = by_cat[cat]["total"]
        diff = abs(actual - expected)
        print(f"    {cat:<30} ${actual:>10,.2f}  (expected ~${expected:,.0f}) {'OK' if diff < 10 else 'CHECK'}")


def main():
    print("Loading Excel data...")
    income_rows, expense_rows = load_excel_data()

    print("\nBuilding sheet data...")
    sheet_data = build_sheet_data(income_rows, expense_rows)
    print(f"Total rows to write: {len(sheet_data)}")

    print("\nAuthenticating with Google Sheets...")
    service = get_sheets_service()

    print("\nClearing and writing data...")
    sheet_id = clear_and_write(service, sheet_data)

    if sheet_id is None:
        print("FAILED: Could not find the Transactions sheet.")
        return

    print("\nApplying formatting...")
    apply_formatting(service, sheet_id, income_rows, expense_rows)

    write_summary(income_rows, expense_rows)

    print("\nDone! Sheet rebuilt successfully.")
    print(f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}")


if __name__ == "__main__":
    main()
